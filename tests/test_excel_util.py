from openpyxl import Workbook, load_workbook
from excel_util import calculate_file_total, create_report
def test_calculate_file_total(tmp_path):
    file_path = tmp_path / "test_sales.xlsx"
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Product"
    sheet["B1"] = "Quantity"
    sheet["C1"] = "Price"
    sheet["A2"] = "Pen"
    sheet["B2"] = 2
    sheet["C2"] = 100
    sheet["A3"] = "Book"
    sheet["B3"] = 3
    sheet["C3"] = 50

    workbook.save(file_path)

    result = calculate_file_total(file_path)

    assert result == 350
def test_create_report(tmp_path):
    report={"file1.xlsx":350
            ,"file2.xlsx":650
    }
    output_file=tmp_path / "report.xlsx"
    create_report(report, output_file)
    assert output_file.exists()
    workbook= load_workbook(output_file)
    sheet = workbook.active
    assert sheet["A2"].value == "file1.xlsx"
    assert sheet["B2"].value == 350
    assert sheet["A4"].value == "TOTAL"
    assert sheet["B4"].value == 1000