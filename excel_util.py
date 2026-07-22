import os
from openpyxl import load_workbook
from openpyxl import Workbook
def calculate_file_total (path):
    workbook=load_workbook(path)
    sheet=workbook.active
    file_total=0
    for row in range (2,sheet.max_row+1):
        quantity=sheet[f"B{row}"].value
        price=sheet[f"C{row}"].value
        if quantity is not None and price is not None:
            file_total+=quantity*price
    return(file_total)
def create_report(report,output_path):
    output=Workbook()
    sheet=output.active
    sheet["A1"]="File"
    sheet["B1"]="Sales"
    row=2
    total_sales=0
    for file,sales in report.items():
        sheet[f"A{row}"]=file
        sheet[f"B{row}"]=sales
        row+=1
        total_sales+=sales
    sheet[f"A{row}"]="TOTAL"
    sheet[f"B{row}"]=total_sales
    output.save(output_path)
def get_excel_files(folder_path):
    if not os.path.exists(folder_path):
        return []
    excel_files=[]
    files=os.listdir(folder_path)
    for file in files:
        if file.lower().endswith(".xlsx"):
            path=os.path.join(folder_path,file)
            excel_files.append(path)
    return excel_files        